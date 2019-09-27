import openpyxl
xl_handle = openpyxl.load_workbook(r'D:\Horace\Documents\workspace\测算参考2.0.xlsx')

get_sheet = xl_handle.worksheets[0]
get_rows = get_sheet.max_row
get_cols = get_sheet.max_column
#151434.4
val_change = get_sheet.cell(9,3).value + 1000
get_sheet.cell(row=9, column=3, value=val_change)
xl_handle.save(r'D:\Horace\Documents\workspace\测算参考2.0.xlsx')
xl_handle.close()

res_xl = openpyxl.load_workbook(r'D:\Horace\Documents\workspace\测算参考2.0.xlsx', data_only=True)
res_sheet = res_xl.worksheets[0]
print("The row of sheet is ",get_rows,"\n","The columns of sheet is ",get_cols,"\n",res_sheet.cell(9,3).value, "\n",res_sheet.cell(26, 7).value)
#result_xl.save(r'D:\Horace\Documents\workspace\测算参考2.0.xlsx')
res_xl.close()
