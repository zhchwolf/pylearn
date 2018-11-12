#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import xdrlib
import sys
import os
import xlrd
import xlsxwriter
import openpyxl
import xlwings
from collections import defaultdict

data = xlrd.open_workbook("BDayShift.xlsx")
table1 = data.sheets()[0]  # 通过索引顺序获取
table2 = data.sheet_by_index(2)  # 通过索引顺序获取
table3 = data.sheet_by_name(u'4-3')  # 通过名称获取

# 获取整行和整列的值（数组）
table1.row_values(3)
table2.col_values(1)
# 获取行数和列数
nrows = table1.nrows
ncols = table1.ncols
#循环行列表数据
for i in range(nrows):
    print (table1.row_values(i))
# 单元格
cell_A1 = table1.cell(0, 0).value
cell_C4 = table1.cell(2, 3).value
#使用行列索引
cell_A1 = table1.row(0)[0].value
cell_A2 = table1.col(1)[0].value
#简单的写入
row = 0
col = 0
# 类型 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
ctype = 1
value = '单元格的值'
xf = 0  # 扩展的格式化
table1.put_cell(row, col, ctype, value, xf)
table1.cell(0, 0)  # 单元格的值'
table1.cell(0, 0).value  # 单元格的值'

# Demo
file_name = 'BDayShift.xlsx'
dir_path=os.curdir
file_path =dir_path + '/' + file_name
print(file_path)
def open_excel(file):
    try:
        data = xlrd.open_workbook(file)
        return data
    except Exception:
        print(str(Exception))
#根据索引获取Excel表格中的数据   参数:file：Excel文件路径     colnameindex：表头列名所在行的所以  ，by_index：表的索引
def excel_table_byindex(file_path,colnameindex=0,by_index=0):
    data = open_excel(file_path)
    table = data.sheets()[by_index]
    nrows = table.nrows #行数
    ncols = table.ncols #列数
    colnames =  table.row_values(colnameindex) #某一行数据
    list =[]
    for rownum in range(1,nrows):
         row = table.row_values(rownum)
         if row:
             app = {}
             for i in range(len(colnames)):
                app[colnames[i]] = row[i]
             list.append(app)
    return list
#根据名称获取Excel表格中的数据   参数:file：Excel文件路径     colnameindex：表头列名所在行的所以  ，by_name：Sheet1名称
def excel_table_byname(file_path,colnameindex=0,by_name=u'4-1'):
    data = open_excel(file_path)
    table = data.sheet_by_name(by_name)
    nrows = table.nrows #行数
    colnames =  table.row_values(colnameindex) #某一行数据
    list =[]
    for rownum in range(1,nrows):
         row = table.row_values(rownum)
         if row:
             app = {}
             for i in range(len(colnames)):
                app[colnames[i]] = row[i]
             list.append(app)
    return list
#根据名称获取Excel表格中指定值的index   参数:file：Excel文件路径
def get_cell_index(file,by_name):
    exceldata = open_excel(file)
    table3 = exceldata.sheet_by_name(by_name)
    nrows = table3.nrows  # 行数
    ncols = table3.ncols  # 列数
    for i in range(0, ncols):
        #print('i:'+ i)
        for j in range(0, nrows):
            #print(j)
            #print(table3.cell(j,i).value)
            if table3.cell(j,i).value == 'Total':
                #print(table3.cell(j, i).value)
                break
    return j

def is_null(value_input):
    if type(value_input) == str:
        return 0
    else:
        return value_input

def check_worktime_sheet(file,sh_name):
    d = defaultdict(list)
    n = get_cell_index(file,sh_name)
    exldata = xlrd.open_workbook(file).sheet_by_name(sh_name)
    sheet1 = exldata.name
    print('sheet name:',sheet1)
    for i in list(range(n))[4:-2]:
        #print(exldata.cell(i,4).value , exldata.cell(i,5).value)
        d[exldata.cell(i,4).value].append(is_null(exldata.cell(i,6).value))
        d[exldata.cell(i,4).value].append(is_null(exldata.cell(i,7).value))
    for k in d.keys():
        if int(sum(d.get(k))) > 10:
            print('工号：',k)
            print('劳动分时：',d.get(k))
            print('总劳动时间：',sum(d.get(k)))

def check_worktime(file):
    sheet_names_list = xlrd.open_workbook(file).sheet_names()
    print(sheet_names_list)
    for s in sheet_names_list:
        #print(s)
        check_worktime_sheet(file,s)

    #for n in tables:

check_worktime(file_path)